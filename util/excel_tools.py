import os
import xlrd
from openpyxl import Workbook


def read_xls_file(file_name, sheet_name):
    """
    :param file_name:
    :param sheet_name:
    :return: list of list ; like [[], []]
    """
    data = xlrd.open_workbook(file_name)
    table = data.sheet_by_name(sheet_name)
    return [table.row_values(i) for i in range(0, table.nrows)]


def write_data_multiple_into_excel(filename, sheetname_list, data_list_list, over_write=False,
                                   out_path='.', not_orver_write_append='(1)'):
    """

    :param filename:   eg:  xxx
    :param sheetname_list:  eg: ["xxx"]
    :param data_list_list:  eg: [  [  ["xxx"],["xxx"] ]  ]
    :param over_write:
    :param out_path:
    :param not_orver_write_append:
    :return:
    """
    if len(sheetname_list) != len(data_list_list):
        return 'sheetname_list and data_list_list not match!'
    wb = Workbook()
    for data_index, sheetname in enumerate(sheetname_list):
        data_list = data_list_list[data_index]
        ws = wb.create_sheet(sheetname, 0)

        for i, data in enumerate(data_list, 1):
            try:
                for j, d in enumerate(data, 1):
                    ws.cell(row=i, column=j).value = d
            except:
                ws.cell(row=i, column=j).value = u"DATA_ERROR"
    if not filename:
        filename = 'new_file'
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    filename_pre = filename[:-5]
    filename_tag = filename[-5:]

    if not over_write:
        while os.path.exists(os.path.join(out_path, filename_pre + filename_tag)):
            filename_pre += not_orver_write_append
    wb.save(os.path.join(out_path, filename_pre + filename_tag))


def example_read_xls_file():
    """
    :return:
    """
    data_list = read_xls_file("test.xlsx",
                              "test")
    for data in data_list:
        print(data)


def example_write_data_multiple_into_excel():
    """
    :return:
    """

    filename = 'test'
    sheetname_list = ['test']
    # sheet test A0 will be 111
    data_list_list = [[['111']]]
    write_data_multiple_into_excel(filename, sheetname_list, data_list_list)


if __name__ == "__main__":
    example_write_data_multiple_into_excel()

    example_read_xls_file()
